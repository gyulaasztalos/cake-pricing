"""Daily automatic price update from the árfigyelő XLSX feed.

Flow (see app/jobs/price_sync.py for the runnable entrypoint):
  download the XLSX -> parse product_id -> price -> for every component that has
  a "Termék azonosító", look it up and, if the price differs, record a temporal
  price change (§3.4: close the open window, insert a new one). Components whose
  id is not found get a price_missing_at marker for the UI warning.

File shape (verified against the live file):
  * one worksheet, header in row 1, data from row 2;
  * column A = Termék azonosító — an ARBITRARY text id, matched EXACTLY. Most are
    zero-padded digit strings ('0000000022989'), but ~5% are chain-prefixed
    ('aldi-10026', 'lidl-...'); everything is read/compared as a string, so both
    work with no numeric assumptions;
  * column I = "Maximum ár" (Hungarian decimal comma, e.g. '499,0000').
  * the SAME product id can appear on several rows (one per store chain) with
    different prices — we AVERAGE them, rounded to whole forint (owner's choice).

Everything raises on failure so the CronJob fails loudly (→ AlertManager).
"""

from __future__ import annotations

import datetime as dt
import warnings
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

import httpx
import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Component, ComponentPrice, PriceSyncState
from app.routers._helpers import decimal_hu

# Column indices (0-based) in the worksheet.
_COL_PRODUCT_ID = 0  # A — Termék azonosító
_COL_PRICE = 8  # I — Maximum ár

# Chains sometimes publish a plainly wrong row (a 1 kg bag of sugar listed at
# 2990 Ft next to 318 and 331). A spread this wide means one of the rows is bad,
# not that the market disagrees.
_IMPLAUSIBLE_SPREAD = Decimal("3")

_DOWNLOAD_TIMEOUT = 120.0


@dataclass(frozen=True)
class PriceChange:
    component_id: int
    name: str
    old_price: Decimal | None
    new_price: Decimal


@dataclass(frozen=True)
class FeedPrice:
    """One product's price as read from the feed, with the evidence behind it."""

    value: Decimal  # median across chains, whole forint
    samples: int
    low: Decimal
    high: Decimal
    reliable: bool


@dataclass(frozen=True)
class UnreliableFeed:
    """A component whose feed rows disagree too much to pick a price from."""

    name: str
    low: Decimal
    high: Decimal


@dataclass
class SyncResult:
    changes: list[PriceChange] = field(default_factory=list)
    missing: list[str] = field(default_factory=list)  # component names not found
    unreliable: list[UnreliableFeed] = field(default_factory=list)  # feed rows disagree
    checked: int = 0  # components with a product_id


def fetch_xlsx(url: str) -> bytes:
    """Download the price file. Raises on any HTTP/transport error."""
    resp = httpx.get(url, timeout=_DOWNLOAD_TIMEOUT, follow_redirects=True, trust_env=False)
    resp.raise_for_status()
    return resp.content


def parse_prices(data: bytes) -> dict[str, FeedPrice]:
    """Map exact Termék azonosító -> the MEDIAN Maximum ár (whole forint).

    Reading is streamed (read_only) so the ~25k-row file stays light. Rows with
    a blank id or an unparseable price are skipped.

    The median, not the mean: a product id appears once per store chain, and a
    single chain publishing nonsense used to drag the average far away from every
    real price. A 1 kg bag of Koronás sugar listed at 318 / 2990 / 331 averaged to
    1213 Ft — a price no shop charged. The median returns 331 and ignores the bad
    row. In one sample of the feed, 7 products had a chain price 3x above the
    median, so this is a recurring property of the source, not a one-off.

    With only TWO rows the median sits midway between them, so it cannot outvote a
    bad one; such products are marked `reliable=False` (106 of them in that same
    sample) and the caller declines to price from them rather than guessing.
    """
    with warnings.catch_warnings():
        # The feed ships without a default style; openpyxl warns but reads fine.
        warnings.simplefilter("ignore", category=UserWarning)
        workbook = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        collected: dict[str, list[Decimal]] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= _COL_PRICE:
                continue
            raw_id = row[_COL_PRODUCT_ID]
            if raw_id is None or not str(raw_id).strip():
                continue
            price = decimal_hu(str(row[_COL_PRICE]) if row[_COL_PRICE] is not None else "")
            if price is None or price < 0:
                continue
            collected.setdefault(str(raw_id).strip(), []).append(price)
    finally:
        workbook.close()

    prices: dict[str, FeedPrice] = {}
    for product_id, values in collected.items():
        ordered = sorted(values)
        mid = len(ordered) // 2
        median = (
            ordered[mid]
            if len(ordered) % 2
            else (ordered[mid - 1] + ordered[mid]) / 2  # even count: midpoint
        )
        low, high = ordered[0], ordered[-1]
        # Two rows that disagree wildly leave nothing to arbitrate between; three or
        # more let the median outvote the odd one out, so those stay usable.
        wide = low > 0 and high / low >= _IMPLAUSIBLE_SPREAD
        prices[product_id] = FeedPrice(
            value=median.quantize(Decimal("1"), rounding=ROUND_HALF_UP),  # whole forint
            samples=len(ordered),
            low=low,
            high=high,
            reliable=not (wide and len(ordered) < 3),
        )
    return prices


def _current_open_price(session: Session, component_id: int) -> ComponentPrice | None:
    """The component's currently-open (unexpired) price row, if any."""
    return session.scalars(
        select(ComponentPrice)
        .where(
            ComponentPrice.component_id == component_id,
            ComponentPrice.expiration_date.is_(None),
        )
        .order_by(ComponentPrice.effective_date.desc())
    ).first()


def _apply_price_change(session: Session, open_row: ComponentPrice, new_price: Decimal) -> None:
    """Temporal price change (§3.4): close the open row and insert a new one at
    the same instant, carrying the existing base_amount and just moving the
    price. Mirrors the manual change_price endpoint."""
    now = dt.datetime.now(dt.UTC)
    # The open row's effective_date came from the DB's now() (transaction time);
    # a small clock skew between the DB and this process could otherwise leave
    # `now` <= effective_date and violate the window CHECK (expiration > effective).
    if open_row.effective_date >= now:
        now = open_row.effective_date + dt.timedelta(milliseconds=1)
    open_row.expiration_date = now
    session.flush()  # apply expiration before insert to satisfy the EXCLUDE constraint
    session.add(
        ComponentPrice(
            component_id=open_row.component_id,
            base_amount=open_row.base_amount,
            base_price=new_price,
            effective_date=now,
        )
    )


def run_sync(session: Session, prices: dict[str, FeedPrice]) -> SyncResult:
    """Reconcile every component that has a product_id against `prices`.

    Does NOT commit — the caller commits once, after (optionally) e-mailing, so a
    mail failure never leaves a half-applied run. `last_success_at` is bumped
    separately by mark_success().
    """
    result = SyncResult()
    now = dt.datetime.now(dt.UTC)
    components = session.scalars(
        select(Component).where(Component.product_id.is_not(None)).order_by(Component.name)
    ).all()
    for comp in components:
        product_id = (comp.product_id or "").strip()
        if not product_id:
            continue
        result.checked += 1
        feed = prices.get(product_id)  # exact-string match (owner's choice)
        if feed is None:
            comp.price_missing_at = now
            result.missing.append(comp.name)
            continue
        comp.price_missing_at = None  # found → clear any earlier warning
        if not feed.reliable:
            # Keep the last known-good price rather than overwrite it with a coin
            # flip between two contradictory rows. Reported so it is not silent.
            result.unreliable.append(UnreliableFeed(comp.name, feed.low, feed.high))
            continue
        new_price = feed.value
        open_row = _current_open_price(session, comp.id)
        old_price = open_row.base_price if open_row else None
        if open_row is None:
            # No price yet: seed one (base_amount 1 → price is per single unit).
            session.add(
                ComponentPrice(component_id=comp.id, base_amount=Decimal("1"), base_price=new_price)
            )
            result.changes.append(PriceChange(comp.id, comp.name, None, new_price))
        elif open_row.base_price != new_price:
            _apply_price_change(session, open_row, new_price)
            result.changes.append(PriceChange(comp.id, comp.name, old_price, new_price))
    return result


def mark_success(session: Session, when: dt.datetime | None = None) -> None:
    """Record a successful run (for the /metrics staleness gauge)."""
    state = session.get(PriceSyncState, 1)
    if state is None:
        state = PriceSyncState(id=1)
        session.add(state)
    state.last_success_at = when or dt.datetime.now(dt.UTC)
